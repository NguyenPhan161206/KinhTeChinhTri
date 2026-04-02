import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def update_salary_with_kpi():
    file_path = 'financial_model_Final.xlsx'
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"LỖI: Không thể mở tệp. Hãy chắc chắn tệp tồn tại. Chi tiết: {e}")
        return

    if 'Luong_theo_ban' in wb.sheetnames:
        del wb['Luong_theo_ban']
    ws = wb.create_sheet('Luong_theo_ban')

    # Định nghĩa cấu trúc lương và thưởng (Triệu VNĐ)
    # Định dạng: {Bộ phận: [(Vị trí, Lương cơ bản, Thưởng KPI)]}
    salary_structure = {
        'Marketing': [
            ('Manager', 40, 10), 
            ('Performance', 25, 5), 
            ('Content', 15, 3), 
            ('Creative', 15, 3)
        ],
        'Sales': [
            ('Team lead', 30, 10), 
            ('Defendence', 20, 5), 
            ('Sales member', 15, 5)
        ],
        'Tech': [
            ('Automation', 35, 7), 
            ('Fullstack', 30, 6), 
            ('Devops', 30, 6)
        ],
        'Ops': [
            ('Ops manager', 30, 5), 
            ('Delivery', 15, 2), 
            ('Support', 15, 2)
        ]
    }

    # Tiêu đề bảng mới
    headers = ['Tháng', 'Bộ phận', 'Vị trí', 'Số lượng', 'Lương cơ bản (Tr)', 'Thưởng KPI (Tr)', 'Tổng lương (Tr)']
    ws.append(headers)

    # Định dạng tiêu đề
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Điền dữ liệu cho 12 tháng
    for m in range(1, 13):
        for dept, roles in salary_structure.items():
            for pos, base, kpi in roles:
                # Tính số lượng nhân sự tăng dần theo tháng
                count = 1
                if pos in ['Sales member', 'Performance', 'Content', 'Fullstack']:
                    count = 2 + (m // 4) # Tăng nhẹ
                
                total = (base + kpi) * count
                ws.append([m, dept, pos, count, base, kpi, total])

    # Tự động căn chỉnh độ rộng cột
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    try:
        wb.save(file_path)
        print("SUCCESS: Đã cập nhật bảng Luong_theo_ban (bao gồm Lương & Thưởng KPI) thành công.")
    except PermissionError:
        print("LỖI: Vui lòng đóng tệp Excel 'financial_model_Final.xlsx' trước khi chạy lệnh này.")

if __name__ == "__main__":
    update_salary_with_kpi()
