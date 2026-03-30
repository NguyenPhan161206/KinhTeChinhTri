import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

file_path = "financial_model.xlsx"
wb = openpyxl.load_workbook(file_path)

# Update Luong_theo_ban to show growth
ws_l = wb["Luong_theo_ban"]
# Find the "Sales Member" role for each month
# Roles are listed for each month. Month 1 is row 2 to 16. Month 2 is 17 to 31.
# Roles data (15 roles):
# Marketing: Manager, Performance, Content, Creative
# Sales: Teamlead, Defence, Member (row 7 in roles list, so index 6)
# Accountant: Accountant
# Tech: Automation, Fullstack, DevOps
# OPS: Quản lý vận hành, Triển khai, Kiểm soát, Hỗ trợ

# Roles start at row 10 in Dau_vao.
# Sales Member is role 7, so it's row 16 in Dau_vao.

# In Luong_theo_ban, Month m starts at row (m-1)*15 + 2
# Sales Member is the 7th role in the list, so its row is (m-1)*15 + 2 + 6 = (m-1)*15 + 8

for m in range(1, 13):
    row_idx = (m - 1) * 15 + 8
    # Ensure it's Sales Member
    role_name = ws_l.cell(row=row_idx, column=3).value
    if role_name == "Member":
        # Update Headcount for this role in Month m
        # Value = Original Count + (m * Monthly Hiring)
        # B6 is Hiring rate
        ws_l.cell(row=row_idx, column=4, value=f"=Dau_vao!$C$16 + ({m} * Assumptions!$B$6)")
    else:
        # Fallback search if index changed
        for i in range((m-1)*15 + 2, m*15 + 2):
            if ws_l.cell(row=i, column=3).value == "Member" and ws_l.cell(row=i, column=2).value == "Sales":
                ws_l.cell(row=i, column=4, value=f"=Dau_vao!$C$16 + ({m} * Assumptions!$B$6)")
                break

# Now that Luong_theo_ban reflects the growth, 
# we can safely set Financial Model column E to link back to Luong_theo_ban.
# This makes the model "Deeply integrated" again.

ws_fm = wb["Financial Model"]
currency_fmt = '#,##0'
for m in range(1, 13):
    r = m + 2
    # Link back to SUMIFS of Luong_theo_ban
    ws_fm.cell(row=r, column=5, value=f"=SUMIFS(Luong_theo_ban!G:G, Luong_theo_ban!A:A, A{r})").number_format = currency_fmt

wb.save(file_path)
print("SUCCESS: Updated Luong_theo_ban headcount and relinked Financial Model.")
