import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def sync_salary_formulas():
    file_path = 'financial_model_Final.xlsx'
    try:
        wb = openpyxl.load_workbook(file_path)
    except: return

    # 1. Cập nhật Luong_theo_ban với công thức
    ws_l = wb['Luong_theo_ban']
    ws_l.delete_rows(2, ws_l.max_row)
    
    salary_structure = {
        'Marketing': [('Manager', 40, 10), ('Performance', 25, 5), ('Content', 15, 3), ('Creative', 15, 3)],
        'Sales': [('Team lead', 30, 10), ('Defendence', 20, 5), ('Sales member', 15, 5)],
        'Tech': [('Automation', 35, 7), ('Fullstack', 30, 6), ('Devops', 30, 6)],
        'Ops': [('Ops manager', 30, 5), ('Delivery', 15, 2), ('Support', 15, 2)]
    }

    current_row = 2
    for m in range(1, 13):
        for dept, roles in salary_structure.items():
            for pos, base, kpi in roles:
                count = 1
                if pos in ['Sales member', 'Performance', 'Content', 'Fullstack']:
                    count = 2 + (m // 4)
                
                # Chèn dữ liệu và công thức: Tổng lương = (E + F) * D
                formula = f"=(E{current_row}+F{current_row})*D{current_row}"
                ws_l.append([m, dept, pos, count, base, kpi, formula])
                current_row += 1

    # 2. Cập nhật KPI_theo_ban (Quỹ lương = SUM của Tổng lương cá nhân)
    ws_k = wb['KPI_theo_ban']
    ws_k.delete_rows(2, ws_k.max_row)
    departments = ['Marketing', 'Sales', 'Tech', 'Ops']
    
    row_k = 2
    for m in range(1, 13):
        for dept in departments:
            # SUMIFS lấy từ Luong_theo_ban
            f_count = f"=SUMIFS(Luong_theo_ban!D:D, Luong_theo_ban!A:A, A{row_k}, Luong_theo_ban!B:B, B{row_k})"
            f_fund = f"=SUMIFS(Luong_theo_ban!G:G, Luong_theo_ban!A:A, A{row_k}, Luong_theo_ban!B:B, B{row_k})"
            ws_k.append([m, dept, f_count, f_fund, None, None, None])
            row_k += 1

    wb.save(file_path)
    print("SUCCESS: Đã đồng bộ quỹ lương bằng công thức Excel.")

if __name__ == "__main__":
    sync_salary_formulas()
