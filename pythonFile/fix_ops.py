import openpyxl

def fix_ops_sync():
    staff_data = {
        1: {'sales': 5, 'mkt': 10, 'tech': 10, 'ops': 4},
        2: {'sales': 6, 'mkt': 10, 'tech': 10, 'ops': 4},
        3: {'sales': 6, 'mkt': 10, 'tech': 10, 'ops': 5},
        4: {'sales': 7, 'mkt': 10, 'tech': 10, 'ops': 5},
        5: {'sales': 7, 'mkt': 11, 'tech': 10, 'ops': 5},
        6: {'sales': 7, 'mkt': 11, 'tech': 11, 'ops': 5},
        7: {'sales': 8, 'mkt': 11, 'tech': 11, 'ops': 5},
        8: {'sales': 8, 'mkt': 12, 'tech': 11, 'ops': 5},
        9: {'sales': 8, 'mkt': 12, 'tech': 12, 'ops': 5},
        10: {'sales': 9, 'mkt': 12, 'tech': 12, 'ops': 5},
        11: {'sales': 9, 'mkt': 13, 'tech': 12, 'ops': 5},
        12: {'sales': 10, 'mkt': 13, 'tech': 12, 'ops': 5}
    }

    file_path = 'financial_model_Final.xlsx'
    wb = openpyxl.load_workbook(file_path)
    ws = wb['Luong_theo_ban']

    for row in range(2, ws.max_row + 1):
        month = ws.cell(row=row, column=1).value
        dept = ws.cell(row=row, column=2).value
        role = ws.cell(row=row, column=3).value
        
        if month in staff_data and dept == 'OPS':
            total_ops = staff_data[month]['ops']
            # Phân bổ lại OPS:
            if 'Quản lý' in str(role): 
                ws.cell(row=row, column=4, value=1)
            elif 'Triển khai' in str(role):
                ws.cell(row=row, column=4, value=2 if total_ops >= 5 else 1)
            elif 'Kiểm soát' in str(role):
                ws.cell(row=row, column=4, value=1)
            elif 'Hỗ trợ' in str(role):
                ws.cell(row=row, column=4, value=1 if total_ops >= 5 else 1)
                # Đảm bảo tổng khớp 4 hoặc 5
                current_sum = 1 + (2 if total_ops >= 5 else 1) + 1 + 1
                if current_sum > total_ops:
                    ws.cell(row=row, column=4, value=0)

    wb.save(file_path)
    print("Fixed OPS staffing in Excel.")

if __name__ == "__main__":
    fix_ops_sync()
