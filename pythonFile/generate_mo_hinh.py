import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_model():
    wb = openpyxl.Workbook()
    
    # --- Sheet 2: Dau vao (Inputs) ---
    ws_input = wb.active
    ws_input.title = "Dau vao"
    
    inputs = [
        ("Thông số", "Giá trị"),
        ("Vốn ban đầu (VND)", 2000000000),
        ("Tỷ lệ tăng trưởng hàng tháng", 0.05),
        ("Chi phí cố định hàng tháng (VND)", 25000000),
        ("Số tháng phân tích", 12)
    ]
    
    for row in inputs:
        ws_input.append(row)
        
    # Formatting Sheet 2
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    white_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws_input["1:1"]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        
    ws_input.column_dimensions['A'].width = 35
    ws_input.column_dimensions['B'].width = 20
    
    ws_input["B2"].number_format = '#,##0'
    ws_input["B3"].number_format = '0%'
    ws_input["B4"].number_format = '#,##0'

    # --- Sheet 1: Mo hinh tai chinh ---
    ws_model = wb.create_sheet("Mo hinh tai chinh", 0)
    
    headers = ["Tháng", "GTDN", "Lãi tháng", "Tổng lương tháng (Nhập tay)", "Chi phí cố định", "TNTT"]
    ws_model.append(headers)
    
    # Formatting Header
    for cell in ws_model["1:1"]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    
    # Input cell highlight style
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Tháng 0 (Row 2)
    ws_model.cell(row=2, column=1, value=0)
    ws_model.cell(row=2, column=2, value="='Dau vao'!$B$2") # GTDN
    ws_model.cell(row=2, column=3, value=0) # Lãi tháng
    ws_model.cell(row=2, column=4, value=0).fill = input_fill # Lương (editable)
    ws_model.cell(row=2, column=5, value=0) # CP cố định
    ws_model.cell(row=2, column=6, value=0) # TNTT tháng 0 = 0
    
    # Tháng 1 - 12 (Rows 3 to 14)
    for month in range(1, 13):
        row_idx = month + 2
        prev_row = row_idx - 1
        
        # A: Tháng
        ws_model.cell(row=row_idx, column=1, value=month)
        
        # B: GTDN_n = GTDN_(n-1) * (1 + Growth)
        ws_model.cell(row=row_idx, column=2, value=f"=B{prev_row}*(1+'Dau vao'!$B$3)")
        
        # C: Lãi tháng_n = GTDN_(n-1) * Growth
        ws_model.cell(row=row_idx, column=3, value=f"=B{prev_row}*'Dau vao'!$B$3")
        
        # D: Lương tháng (Manual Input)
        salary_cell = ws_model.cell(row=row_idx, column=4, value=0)
        salary_cell.fill = input_fill
        salary_cell.border = border
        
        # E: Chi phí cố định
        ws_model.cell(row=row_idx, column=5, value="='Dau vao'!$B$4")
        
        # F: TNTT_n = TNTT_(n-1) + Lãi tháng_n - Lương_n - CP cố định
        ws_model.cell(row=row_idx, column=6, value=f"=F{prev_row}+C{row_idx}-D{row_idx}-E{row_idx}")

    # Final Formatting for Sheet 1
    col_widths = [10, 20, 20, 25, 20, 25]
    for i, width in enumerate(col_widths, 1):
        ws_model.column_dimensions[get_column_letter(i)].width = width
        
    for row in ws_model.iter_rows(min_row=2, max_row=14, min_col=2, max_col=6):
        for cell in row:
            cell.number_format = '#,##0'
            if cell.column != 4: # Add thin border to non-input cells for clarity
                cell.border = border

    wb.save("mo_hinh_tai_chinh.xlsx")

if __name__ == "__main__":
    generate_excel_model()
