import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

file_path = "financial_model.xlsx"

# Create a fresh workbook to fix corruption
wb = openpyxl.Workbook()

# --- Styles ---
bold_font = Font(bold=True)
white_bold_font = Font(bold=True, color="FFFFFF")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
blue_fill = PatternFill(patternType='solid', fgColor="4F81BD")
header_fill = PatternFill(patternType='solid', fgColor="D9D9D9")
input_fill = PatternFill(patternType='solid', fgColor="FFF2CC")
currency_fmt = '#,##0'

# 1. SHEET: Assumptions (Restore original)
ws_assump = wb.active
ws_assump.title = "Assumptions"
assumptions_data = [
    ("Tham số", "Giá trị"),
    ("Vốn đầu tư ban đầu (VND)", 2000000000),
    ("Tỷ lệ tăng trưởng hàng tháng", 0.05),
    ("Lương bình quân (VND/người)", 8000000),
    ("Số nhân sự ban đầu", 30),
    ("Tuyển dụng hàng tháng", 1)
]
for row in assumptions_data:
    ws_assump.append(row)

# Formatting Assumptions
for cell in ws_assump["1:1"]:
    cell.font = white_bold_font
    cell.fill = blue_fill
ws_assump.column_dimensions['A'].width = 35
ws_assump.column_dimensions['B'].width = 20
ws_assump["B3"].number_format = '0%'
ws_assump["B2"].number_format = currency_fmt
ws_assump["B4"].number_format = currency_fmt

# 2. SHEET: Dau_vao (New structure)
ws_dv = wb.create_sheet("Dau_vao")
ws_dv["A1"] = "SECTION 1: THÔNG SỐ CỐ ĐỊNH (LINKED)"
ws_dv["A1"].font = bold_font
dv_params = [
    ("Vốn đầu tư ban đầu", "=Assumptions!$B$2"),
    ("Tỷ lệ tăng trưởng", "=Assumptions!$B$3"),
    ("Lương bình quân (Avg Salary)", "=Assumptions!$B$4"),
    ("Tỷ lệ Lương cứng (Base %)", 0.7),
    ("Tỷ lệ Thưởng KPI (KPI %)", 0.3)
]
for i, (label, val) in enumerate(dv_params, 2):
    ws_dv.cell(row=i, column=1, value=label).border = border
    cell_b = ws_dv.cell(row=i, column=2, value=val)
    cell_b.border = border
    if "Tỷ lệ" in label:
        cell_b.number_format = '0%'
        if i >= 5: cell_b.fill = input_fill
    else:
        cell_b.number_format = currency_fmt

ws_dv["A8"] = "SECTION 2-4: CƠ CẤU NHÂN SỰ & HỆ SỐ LƯƠNG"
ws_dv["A8"].font = bold_font
headers_dv = ["Ban", "Role", "Số người", "Lương cơ bản", "Thưởng KPI", "Hệ số (Coef)"]
for c, h in enumerate(headers_dv, 1):
    cell = ws_dv.cell(row=9, column=c, value=h)
    cell.font = bold_font
    cell.fill = header_fill
    cell.border = border

roles_data = [
    ("Marketing", "Manager", 1, 2.5), ("Marketing", "Performance", 2, 1.5),
    ("Marketing", "Content", 2, 1.2), ("Marketing", "Creative", 1, 1.3),
    ("Sales", "Teamlead", 1, 2.0), ("Sales", "Defence", 2, 1.2), ("Sales", "Member", 5, 1.0),
    ("Accountant", "Accountant", 1, 1.5),
    ("Tech", "Automation", 1, 2.0), ("Tech", "Fullstack", 2, 2.2), ("Tech", "DevOps", 1, 2.5),
    ("OPS", "Quản lý vận hành", 1, 2.2), ("OPS", "Triển khai / Delivery", 2, 1.2),
    ("OPS", "Kiểm soát & tối ưu", 1, 1.5), ("OPS", "Hỗ trợ vận hành", 2, 1.0)
]

for i, (ban, role, count, coef) in enumerate(roles_data):
    r = i + 10
    ws_dv.cell(row=r, column=1, value=ban).border = border
    ws_dv.cell(row=r, column=2, value=role).border = border
    ws_dv.cell(row=r, column=3, value=count).fill = input_fill
    ws_dv.cell(row=r, column=3).border = border
    ws_dv.cell(row=r, column=6, value=coef).fill = input_fill
    ws_dv.cell(row=r, column=6).border = border
    ws_dv.cell(row=r, column=4, value=f"=$B$4 * F{r} * $B$5").number_format = currency_fmt
    ws_dv.cell(row=r, column=5, value=f"=$B$4 * F{r} * $B$6").number_format = currency_fmt
    ws_dv.cell(row=r, column=4).border = border
    ws_dv.cell(row=r, column=5).border = border

ws_dv.column_dimensions['A'].width = 30
ws_dv.column_dimensions['B'].width = 25

# 3. SHEET: Luong_theo_ban
ws_l = wb.create_sheet("Luong_theo_ban")
l_headers = ["Tháng", "Ban", "Role", "Số người", "Lương cơ bản", "Thưởng KPI", "Tổng lương role", "Tổng lương ban"]
for c, h in enumerate(l_headers, 1):
    cell = ws_l.cell(row=1, column=c, value=h)
    cell.font = bold_font
    cell.fill = header_fill
    cell.border = border

row_idx = 2
for m in range(1, 13):
    for r_idx in range(len(roles_data)):
        dv_row = r_idx + 10
        ws_l.cell(row=row_idx, column=1, value=m)
        ws_l.cell(row=row_idx, column=2, value=roles_data[r_idx][0])
        ws_l.cell(row=row_idx, column=3, value=roles_data[r_idx][1])
        ws_l.cell(row=row_idx, column=4, value=f"=Dau_vao!$C${dv_row}").fill = input_fill
        ws_l.cell(row=row_idx, column=5, value=f"=Dau_vao!$D${dv_row}").number_format = currency_fmt
        ws_l.cell(row=row_idx, column=6, value=f"=Dau_vao!$E${dv_row}").number_format = currency_fmt
        ws_l.cell(row=row_idx, column=7, value=f"=D{row_idx}*(E{row_idx}+F{row_idx})").number_format = currency_fmt
        ws_l.cell(row=row_idx, column=8, value=f"=SUMIFS(G:G, A:A, A{row_idx}, B:B, B{row_idx})").number_format = currency_fmt
        for c in range(1, 9): ws_l.cell(row=row_idx, column=c).border = border
        row_idx += 1

# 4. SHEET: KPI_theo_ban
ws_k = wb.create_sheet("KPI_theo_ban")
k_headers = ["Tháng", "Ban", "Tổng số người", "Ngân sách lương", "Khối lượng công việc", "KPI chung", "Ghi chú"]
for c, h in enumerate(k_headers, 1):
    cell = ws_k.cell(row=1, column=c, value=h)
    cell.font = bold_font
    cell.fill = header_fill
    cell.border = border

depts = ["Marketing", "Sales", "Accountant", "Tech", "OPS"]
r_k = 2
for m in range(1, 13):
    for dept in depts:
        ws_k.cell(row=r_k, column=1, value=m)
        ws_k.cell(row=r_k, column=2, value=dept)
        ws_k.cell(row=r_k, column=3, value=f"=SUMIFS(Luong_theo_ban!D:D, Luong_theo_ban!A:A, A{r_k}, Luong_theo_ban!B:B, B{r_k})")
        ws_k.cell(row=r_k, column=4, value=f"=SUMIFS(Luong_theo_ban!G:G, Luong_theo_ban!A:A, A{r_k}, Luong_theo_ban!B:B, B{r_k})").number_format = currency_fmt
        for c in range(5, 8): ws_k.cell(row=r_k, column=c).fill = input_fill
        for c in range(1, 8): ws_k.cell(row=r_k, column=c).border = border
        r_k += 1

# 5. SHEET: Financial Model (Restore & Link)
ws_fm = wb.create_sheet("Financial Model", 0)
fm_headers = ["Tháng", "Số nhân sự", "GTDN", "Lãi tháng", "Tổng lương", "TNTT (Dư vốn)"]
for c, h in enumerate(fm_headers, 1):
    cell = ws_fm.cell(row=1, column=c, value=h)
    cell.font = white_bold_font
    cell.fill = blue_fill
    cell.alignment = Alignment(horizontal="center")

# Row 2: Month 0
ws_fm.cell(row=2, column=1, value=0)
ws_fm.cell(row=2, column=2, value="=Assumptions!$B$5")
ws_fm.cell(row=2, column=3, value="=Assumptions!$B$2")
ws_fm.cell(row=2, column=4, value=0)
ws_fm.cell(row=2, column=5, value=0)
ws_fm.cell(row=2, column=6, value="=C2") # TNTT initial

for m in range(1, 13):
    r = m + 2
    ws_fm.cell(row=r, column=1, value=m)
    ws_fm.cell(row=r, column=2, value=f"=B{r-1}+Assumptions!$B$6")
    ws_fm.cell(row=r, column=3, value=f"=C{r-1}*(1+Assumptions!$B$3)")
    ws_fm.cell(row=r, column=5, value=f"=SUMIFS(Luong_theo_ban!G:G, Luong_theo_ban!A:A, A{r})").number_format = currency_fmt
    ws_fm.cell(row=r, column=4, value=f"=C{r}-E{r}").number_format = currency_fmt
    ws_fm.cell(row=r, column=6, value=f"=F{r-1}+D{r}").number_format = currency_fmt

for r in range(1, 15):
    for c in range(1, 7):
        ws_fm.cell(row=r, column=c).border = border
        if r > 1 and c >= 3: ws_fm.cell(row=r, column=c).number_format = currency_fmt

# Freeze & Column widths
ws_fm.column_dimensions['C'].width = 20
ws_fm.column_dimensions['D'].width = 20
ws_fm.column_dimensions['E'].width = 20
ws_fm.column_dimensions['F'].width = 20

wb.save(file_path)
print("SUCCESS: Full reconstruction and link completed.")
