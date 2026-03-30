import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

file_path = "financial_model.xlsx"
wb = openpyxl.load_workbook(file_path)

# 1. Update Assumptions
ws_assump = wb["Assumptions"]
# Find "Lương bình quân" row
salary_row = None
for i in range(1, 10):
    if ws_assump.cell(row=i, column=1).value and "Lương bình quân" in ws_assump.cell(row=i, column=1).value:
        salary_row = i
        break

if salary_row:
    ws_assump.cell(row=salary_row, column=2, value=15000000)
    print(f"Updated Average Salary to 15,000,000 in Assumptions row {salary_row}")

# 2. Update Financial Model Formulas
ws_fm = wb["Financial Model"]
currency_fmt = '#,##0'

# Month 0: Salary is usually 0 as it's the start, but let's check
# Headcount is in Column B, Salary in Column E
# Headcount in Month 0 (row 2)
# Headcount in Month 1 (row 3)

for r in range(3, 15): # Months 1 to 12
    # Update Total Salary (Column E)
    # Total Salary = Headcount (B) * Average Salary (Assumptions!$B$4)
    # This ensures that for every +1 in B, E increases by Assumptions!$B$4 (15M)
    ws_fm.cell(row=r, column=5, value=f"=B{r}*Assumptions!$B$4").number_format = currency_fmt
    
    # Recalculate Lãi tháng (Column D) = GTDN (C) - Tổng lương (E)
    ws_fm.cell(row=r, column=4, value=f"=C{r}-E{r}").number_format = currency_fmt
    
    # Recalculate TNTT / Số dư vốn (Column F) = Prev + Lãi tháng
    ws_fm.cell(row=r, column=6, value=f"=F{r-1}+D{r}").number_format = currency_fmt

# 3. Optional: If the user wants to keep Luong_theo_ban consistent, 
# we should ideally make it grow too. 
# But for now, we satisfy the primary request in the main model.

wb.save(file_path)
print("SUCCESS: Updated total salary logic to increase by 15M per additional person.")
