import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

file_path = "financial_model.xlsx"
wb = openpyxl.load_workbook(file_path)

# 1. Sheet Assumptions: Đảm bảo các tham số đúng yêu cầu
ws_assump = wb["Assumptions"]
ws_assump["A2"] = "Vốn đầu tư ban đầu (VND)"
ws_assump["B2"] = 2000000000
ws_assump["A3"] = "Tỷ lệ tăng trưởng hàng tháng"
ws_assump["B3"] = 0.05
ws_assump["A4"] = "Lương bình quân (VND/người)"
ws_assump["B4"] = 15000000
ws_assump["A5"] = "Số nhân sự ban đầu"
ws_assump["B5"] = 30
ws_assump["A6"] = "Tuyển mới hàng tháng"
ws_assump["B6"] = 1

# 2. Sheet Financial Model: Áp dụng công thức chuẩn
ws_fm = wb["Financial Model"]
# Headers: Tháng (A), Số nhân sự (B), GTDN (C), Lãi tháng (D), Tổng lương (E), TNTT (F)
currency_fmt = '#,##0'

# Tháng 0
ws_fm["A2"] = 0
ws_fm["B2"] = "=Assumptions!$B$5"
ws_fm["C2"] = "=Assumptions!$B$2"
ws_fm["D2"] = 0
ws_fm["E2"] = 0
ws_fm["F2"] = "=C2" # TNTT_0 = GTDN_0

for m in range(1, 13):
    r = m + 2
    # Tháng
    ws_fm.cell(row=r, column=1, value=m)
    # Số nhân sự: N_n = N_(n-1) + Tuyển mới
    ws_fm.cell(row=r, column=2, value=f"=B{r-1}+Assumptions!$B$6")
    # GTDN: GTDN_n = GTDN_(n-1) * 1.05
    ws_fm.cell(row=r, column=3, value=f"=C{r-1}*(1+Assumptions!$B$3)")
    # Lãi tháng: Lai_n = 5% * GTDN_(n-1)
    ws_fm.cell(row=r, column=4, value=f"=Assumptions!$B$3*C{r-1}")
    # Lương tháng: Luong_n = Số nhân sự * Lương bình quân
    ws_fm.cell(row=r, column=5, value=f"=B{r}*Assumptions!$B$4")
    # TNTT: TNTT_n = TNTT_(n-1) + Lai_n - Luong_n
    ws_fm.cell(row=r, column=6, value=f"=F{r-1}+D{r}-E{r}")

wb.save("financial_model_v2.xlsx")
print("SUCCESS: Cấu trúc logic 2 sheet đầu đã được chuẩn hóa trong file financial_model_v2.xlsx.")
