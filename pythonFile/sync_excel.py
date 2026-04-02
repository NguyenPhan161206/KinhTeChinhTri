import openpyxl

def update_excel():
    staff_data = {
        1: {'sales': 5, 'mkt': 10, 'tech': 10, 'ops': 4},
        2: {'sales': 6, 'mkt': 10, 'tech': 10, 'ops': 4},
        3: {'sales': 6, 'mkt': 10, 'tech': 10, 'ops': 5},
        4: {'sales': 7, 'mkt': 10, 'tech': 10, 'ops': 5},
        5: {'sales': 7, 'mkt': 11, 'tech': 10, 'ops': 5},
        6: {'sales': 7, 'mkt': 11, 'tech': 11, 'ops': 5},
        7: {'sales': 8, 'mkt': 11, 'tech': 11, 'ops': 5},
        8: {'sales': 8, 'mkt': 12, 'tech': 11, 'ops': 5},
        9: {'sales': 8, 'mkt': 12, 'tech': 12, 'ops': 5},
        10: {'sales': 9, 'mkt': 12, 'tech': 12, 'ops': 5},
        11: {'sales': 9, 'mkt': 13, 'tech': 12, 'ops': 5},
        12: {'sales': 10, 'mkt': 13, 'tech': 12, 'ops': 5}
    }

    file_path = 'financial_model_Final.xlsx'
    wb = openpyxl.load_workbook(file_path)
    
    # 1. Update Luong_theo_ban
    if 'Luong_theo_ban' in wb.sheetnames:
        ws = wb['Luong_theo_ban']
        # Duyệt qua các hàng từ hàng 2 trở đi
        for row in range(2, ws.max_row + 1):
            month_val = ws.cell(row=row, column=1).value
            dept_name = ws.cell(row=row, column=2).value
            role_name = ws.cell(row=row, column=3).value
            
            if month_val in staff_data:
                data = staff_data[month_val]
                # Phân bổ logic:
                if dept_name == 'Marketing':
                    # Có 4 roles: Manager, Performance, Content, Creative
                    total = data['mkt']
                    if 'Manager' in role_name: ws.cell(row=row, column=4, value=1)
                    elif 'Performance' in role_name: ws.cell(row=row, column=4, value=total // 3)
                    elif 'Content' in role_name: ws.cell(row=row, column=4, value=total // 3)
                    elif 'Creative' in role_name: ws.cell(row=row, column=4, value=total - 1 - (total // 3)*2)
                
                elif dept_name == 'Sales':
                    # Có 3 roles: Teamlead, Defence, Member
                    total = data['sales']
                    if 'Teamlead' in role_name: ws.cell(row=row, column=4, value=1)
                    elif 'Defence' in role_name: ws.cell(row=row, column=4, value=1)
                    elif 'Member' in role_name: ws.cell(row=row, column=4, value=total - 2)
                
                elif dept_name == 'Tech':
                    # Có 3 roles: Automation, Fullstack, DevOps
                    total = data['tech']
                    if 'Automation' in role_name: ws.cell(row=row, column=4, value=total // 3)
                    elif 'Fullstack' in role_name: ws.cell(row=row, column=4, value=total // 3)
                    elif 'DevOps' in role_name: ws.cell(row=row, column=4, value=total - (total // 3)*2)
                
                elif dept_name == 'OPS':
                    # Có 4 roles
                    total = data['ops']
                    if 'Quản lý' in role_name: ws.cell(row=row, column=4, value=1)
                    else: ws.cell(row=row, column=4, value=1 if total > 3 else 0)

    # 2. Update Dau_vao (Tháng 1)
    if 'Dau_vao' in wb.sheetnames:
        ws_dv = wb['Dau_vao']
        # Cập nhật cột C (Số người) cho các vị trí
        mapping = {
            'MarketingManager': 1, 'MarketingPerformance': 4, 'MarketingContent': 3, 'MarketingCreative': 2,
            'SalesTeamlead': 1, 'SalesDefence': 1, 'SalesMember': 3,
            'TechAutomation': 3, 'TechFullstack': 4, 'TechDevOps': 3,
            'OPSQuản lý': 1, 'OPSTriển khai': 1, 'OPSKiểm soát': 1, 'OPSHỗ trợ': 1
        }
        for row in range(10, 24):
            dept = str(ws_dv.cell(row=row, column=1).value)
            role = str(ws_dv.cell(row=row, column=2).value)
            key = dept + role[:10].strip()
            for k, v in mapping.items():
                if dept in k and (role[:5] in k or k[:5] in role):
                    ws_dv.cell(row=row, column=3, value=v)

    wb.save(file_path)
    print("Successfully updated Excel file.")

if __name__ == "__main__":
    update_excel()
