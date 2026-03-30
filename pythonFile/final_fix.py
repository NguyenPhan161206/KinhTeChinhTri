import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

file_path = "financial_model.xlsx"
wb = openpyxl.load_workbook(file_path)

def setup_sheet(name):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)

# Styles
bold = Font(bold=True)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(patternType='solid', fgColor="D9D9D9")
input_fill = PatternFill(patternType='solid', fgColor="FFF2CC")
currency_fmt = '#,##0'

# 1. Update SHEET: "Dau_vao"
ws_dv = setup_sheet("Dau_vao")
ws_dv.column_dimensions['A'].width = 30
ws_dv.column_dimensions['B'].width = 20
ws_dv.column_dimensions['F'].width = 15

# Section 1
ws_dv["A1"] = "SECTION 1: THÔNG SỐ CỐ ĐỊNH (LINKED)"
ws_dv["A1"].font = bold
params = [
    ("Vốn đầu tư ban đầu", "=Assumptions!$B$2"),
    ("Tỷ lệ tăng trưởng", "=Assumptions!$B$3"),
    ("Lương bình quân (Avg Salary)", "=Assumptions!$B$4"),
    ("Tỷ lệ Lương cứng (Base %)", 0.7),
    ("Tỷ lệ Thưởng KPI (KPI %)", 0.3)
]
for i, (label, formula) in enumerate(params, 2):
    ws_dv.cell(row=i, column=1, value=label).border = border
    cell_b = ws_dv.cell(row=i, column=2, value=formula)
    cell_b.border = border
    if "Tỷ lệ" in label:
        cell_b.number_format = '0%'
        if i >= 5: cell_b.fill = input_fill
    else:
        cell_b.number_format = currency_fmt

# Roles
ws_dv["A8"] = "SECTION 2, 3, 4: CƠ CẤU NHÂN SỰ & CÔNG THỨC LƯƠNG"
ws_dv["A8"].font = bold
headers = ["Ban", "Role", "Số người ban đầu", "Lương cơ bản", "Thưởng KPI", "Hệ số (Coef)"]
for c, h in enumerate(headers, 1):
    cell = ws_dv.cell(row=9, column=c, value=h)
    cell.font = bold
    cell.fill = header_fill
    cell.border = border

roles_data = [
    ("Marketing", "Manager", 1, 2.5),
    ("Marketing", "Performance", 2, 1.5),
    ("Marketing", "Content", 2, 1.2),
    ("Marketing", "Creative", 1, 1.3),
    ("Sales", "Teamlead", 1, 2.0),
    ("Sales", "Defence", 2, 1.2),
    ("Sales", "Member", 5, 1.0),
    ("Accountant", "Accountant", 1, 1.5),
    ("Tech", "Automation", 1, 2.0),
    ("Tech", "Fullstack", 2, 2.2),
    ("Tech", "DevOps", 1, 2.5),
    ("OPS", "Quản lý vận hành", 1, 2.2),
    ("OPS", "Triển khai / Delivery", 2, 1.2),
    ("OPS", "Kiểm soát & tối ưu", 1, 1.5),
    ("OPS", "Hỗ trợ vận hành", 2, 1.0)
]

for r_off, (ban, role, count, coef) in enumerate(roles_data):
    r = r_off + 10
    ws_dv.cell(row=r, column=1, value=ban).border = border
    ws_dv.cell(row=r, column=2, value=role).border = border
    cell_c = ws_dv.cell(row=r, column=3, value=count)
    cell_c.fill = input_fill
    cell_c.border = border
    cell_f = ws_dv.cell(row=r, column=6, value=coef)
    cell_f.fill = input_fill
    cell_f.border = border
    ws_dv.cell(row=r, column=4, value=f"=$B$4 * F{r} * $B$5").number_format = currency_fmt
    ws_dv.cell(row=r, column=4).border = border
    ws_dv.cell(row=r, column=5, value=f"=$B$4 * F{r} * $B$6").number_format = currency_fmt
    ws_dv.cell(row=r, column=5).border = border

# 2. Luong_theo_ban
ws_l = setup_sheet("Luong_theo_ban")
l_headers = ["Tháng", "Ban", "Role", "Số người", "Lương cơ bản", "Thưởng KPI", "Tổng lương role", "Tổng lương ban"]
for c, h in enumerate(l_headers, 1):
    cell = ws_l.cell(row=1, column=c, value=h)
    cell.font = bold
    cell.fill = header_fill
    cell.border = border

row_idx = 2
for m in range(1, 13):
    for r_idx in range(len(roles_data)):
        dv_row = r_idx + 10
        ws_l.cell(row=row_idx, column=1, value=m)
        ws_l.cell(row=row_idx, column=2, value=roles_data[r_idx][0])
        ws_l.cell(row=row_idx, column=3, value=roles_data[r_idx][1])
        cell_4 = ws_l.cell(row=row_idx, column=4, value=f"=Dau_vao!$C${dv_row}")
        cell_4.fill = input_fill
        ws_l.cell(row=row_idx, column=5, value=f"=Dau_vao!$D${dv_row}").number_format = currency_fmt
        ws_l.cell(row=row_idx, column=6, value=f"=Dau_vao!$E${dv_row}").number_format = currency_fmt
        ws_l.cell(row=row_idx, column=7, value=f"=D{row_idx}*(E{row_idx}+F{row_idx})").number_format = currency_fmt
        ws_l.cell(row=row_idx, column=8, value=f"=SUMIFS(G:G, A:A, A{row_idx}, B:B, B{row_idx})").number_format = currency_fmt
        for c in range(1, 9):
            ws_l.cell(row=row_idx, column=c).border = border
        row_idx += 1

# 3. KPI_theo_ban
ws_k = setup_sheet("KPI_theo_ban")
k_headers = ["Tháng", "Ban", "Tổng số người", "Ngân sách lương", "Khối lượng công việc tối thiểu", "KPI chung", "Ghi chú"]
for c, h in enumerate(k_headers, 1):
    cell = ws_k.cell(row=1, column=c, value=h)
    cell.font = bold
    cell.fill = header_fill
    cell.border = border

depts_list = ["Marketing", "Sales", "Accountant", "Tech", "OPS"]
for m in range(1, 13):
    for dept in depts_list:
        r_k = ws_k.max_row + 1
        ws_k.cell(row=r_k, column=1, value=m)
        ws_k.cell(row=r_k, column=2, value=dept)
        ws_k.cell(row=r_k, column=3, value=f"=SUMIFS(Luong_theo_ban!D:D, Luong_theo_ban!A:A, A{r_k}, Luong_theo_ban!B:B, B{r_k})")
        ws_k.cell(row=r_k, column=4, value=f"=SUMIFS(Luong_theo_ban!G:G, Luong_theo_ban!A:A, A{r_k}, Luong_theo_ban!B:B, B{r_k})").number_format = currency_fmt
        for c in range(5, 8):
            cell_in = ws_k.cell(row=r_k, column=c)
            cell_in.fill = input_fill
        for c in range(1, 8):
            ws_k.cell(row=r_k, column=c).border = border

# 4. Financial Model Update
ws_fm = wb["Financial Model"]
for m in range(1, 13):
    r = m + 2
    ws_fm.cell(row=r, column=5, value=f"=SUMIFS(Luong_theo_ban!G:G, Luong_theo_ban!A:A, A{r})").number_format = currency_fmt
    ws_fm.cell(row=r, column=4, value=f"=C{r}-E{r}").number_format = currency_fmt

ws_fm.cell(row=1, column=6, value="Số dư vốn").font = bold
ws_fm.cell(row=2, column=6, value="=Assumptions!$B$2").number_format = currency_fmt
for m in range(1, 13):
    r = m + 2
    ws_fm.cell(row=r, column=6, value=f"=F{r-1}+D{r}").number_format = currency_fmt

wb.save(file_path)
print("SUCCESS: Unified update completed.")
