import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

file_path = "financial_model.xlsx"
wb = openpyxl.load_workbook(file_path)

# 1. Update "Dau_vao" with coefficients and formulas
ws_dv = wb["Dau_vao"]

# Add coefficients for each role to drive the mapping from Average Salary
# Coefficient: how many times the average salary this role gets
role_coefficients = {
    "Manager": 2.5,
    "Performance": 1.5,
    "Content": 1.2,
    "Creative": 1.3,
    "Teamlead": 2.0,
    "Defence": 1.2,
    "Member": 1.0,
    "Accountant": 1.5,
    "Automation": 2.0,
    "Fullstack": 2.2,
    "DevOps": 2.5,
    "Quản lý vận hành": 2.2,
    "Triển khai / Delivery": 1.2,
    "Kiểm soát & tối ưu": 1.5,
    "Hỗ trợ vận hành": 1.0
}

# Add a header for Coefficient if not exists (we'll put it in a hidden column or just use it in formula)
# Let's put it in Column F and name it "Hệ số"
ws_dv.cell(row=7, column=6, value="Hệ số").font = Font(bold=True)
ws_dv.cell(row=7, column=6).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
ws_dv.cell(row=7, column=6).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# We'll use a 70/30 split as requested/common for "divide the financial"
split_base = 0.7
split_kpi = 0.3

for row in range(8, 23):
    role_name = ws_dv.cell(row=row, column=2).value
    coef = role_coefficients.get(role_name, 1.0)
    
    # Store coefficient in Col F
    ws_dv.cell(row=row, column=6, value=coef).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws_dv.cell(row=row, column=6).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Update formulas for Lương cơ bản (D) and Thưởng KPI (E)
    # Reference Assumptions!$B$4 (Average salary)
    ws_dv.cell(row=row, column=4, value=f"=Assumptions!$B$4 * F{row} * {split_base}").number_format = '#,##0'
    ws_dv.cell(row=row, column=5, value=f"=Assumptions!$B$4 * F{row} * {split_kpi}").number_format = '#,##0'

# 2. Update "Luong_theo_ban"
# It already has references to Dau_vao!D and E, so we mainly need to ensure the logic flows.
ws_l = wb["Luong_theo_ban"]
# The formulas in Luong_theo_ban are already like =Dau_vao!$D$8, so they will update automatically.
# However, I'll refresh them to be sure.
num_roles = len(role_coefficients)
row_idx = 2
for m in range(1, 13):
    for r_idx in range(num_roles):
        # Base salary reference from Dau_vao
        ws_l.cell(row=row_idx, column=5, value=f"=Dau_vao!$D${r_idx+8}")
        # KPI reference from Dau_vao
        ws_l.cell(row=row_idx, column=6, value=f"=Dau_vao!$E${r_idx+8}")
        row_idx += 1

# 3. Update "Financial Model" to use the calculated totals for "maintenance"
ws_fm = wb["Financial Model"]
# Month 1 is row 3. 
# Total Salary in Month 1 should be the sum of "Tổng lương role" for Month 1 in Luong_theo_ban
for m in range(1, 13):
    row_fm = m + 2
    # Sum of column G (Tổng lương role) in Luong_theo_ban where column A (Tháng) == m
    ws_fm.cell(row=row_fm, column=5, value=f"=SUMIFS(Luong_theo_ban!G:G, Luong_theo_ban!A:A, A{row_fm})").number_format = '#,##0'

# Update Profit/Lãi tháng logic to be realistic: Lãi = GTDN - Tổng lương
# GTDN is Revenue.
for m in range(1, 13):
    row_fm = m + 2
    ws_fm.cell(row=row_fm, column=4, value=f"=C{row_fm}-E{row_fm}").number_format = '#,##0'

# Set GTDN Month 0 to something small if it's currently 2B (Capital)
# Actually I'll leave it as is but ensure the growth is applied correctly.
# The user wants to "maintain the business", so I'll add a "Balance" column to track Capital.
ws_fm.cell(row=1, column=6, value="Số dư vốn").font = Font(bold=True)
ws_fm.cell(row=2, column=6, value="=Assumptions!$B$2") # Initial Capital
for m in range(1, 13):
    row_fm = m + 2
    # Balance = Previous Balance + Profit
    ws_fm.cell(row=row_fm, column=6, value=f"=F{row_fm-1}+D{row_fm}").number_format = '#,##0'

wb.save(file_path)
print("Updated financial_model.xlsx with dynamic salary mapping and business maintenance tracking.")
