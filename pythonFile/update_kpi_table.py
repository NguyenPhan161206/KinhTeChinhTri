import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def update_kpi_sheet():
    file_path = 'financial_model_Final.xlsx'
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"LỖI: Không thể mở tệp. {e}")
        return

    # Chuẩn bị sheet KPI_theo_ban
    if 'KPI_theo_ban' in wb.sheetnames:
        del wb['KPI_theo_ban']
    ws = wb.create_sheet('KPI_theo_ban')

    # Tiêu đề bảng
    headers = ['Tháng', 'Ban', 'Tổng người', 'Quỹ lương (Tr)', 'Công việc', 'KPI', 'Ghi chú']
    ws.append(headers)

    # Định dạng tiêu đề
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Danh sách các ban (khớp chính xác với Luong_theo_ban)
    departments = ['Marketing', 'Sales', 'Tech', 'Ops']

    # Điền dữ liệu cho 12 tháng
    row_idx = 2
    for m in range(1, 13):
        for dept in departments:
            # Công thức SUMIFS:
            # Tổng người (Cột C) = SUMIFS(Luong_theo_ban!D:D, Luong_theo_ban!A:A, A{row}, Luong_theo_ban!B:B, B{row})
            # Quỹ lương (Cột D) = SUMIFS(Luong_theo_ban!G:G, Luong_theo_ban!A:A, A{row}, Luong_theo_ban!B:B, B{row})
            
            formula_count = f"=SUMIFS(Luong_theo_ban!D:D, Luong_theo_ban!A:A, A{row_idx}, Luong_theo_ban!B:B, B{row_idx})"
            formula_salary = f"=SUMIFS(Luong_theo_ban!G:G, Luong_theo_ban!A:A, A{row_idx}, Luong_theo_ban!B:B, B{row_idx})"
            
            ws.append([m, dept, formula_count, formula_salary, None, None, None])
            
            # Định dạng hàng dữ liệu
            for cell in ws[row_idx]:
                cell.alignment = Alignment(horizontal='left')
            
            row_idx += 1

    # Tự động căn chỉnh độ rộng cột
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 5

    try:
        wb.save(file_path)
        print("SUCCESS: Đã chuẩn hóa và liên kết bảng KPI_theo_ban thành công.")
    except PermissionError:
        print("LỖI: Vui lòng đóng tệp Excel trước khi chạy lệnh này.")

if __name__ == "__main__":
    update_kpi_sheet()
