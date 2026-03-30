import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

file_path = "financial_model.xlsx"
wb = openpyxl.load_workbook(file_path)

def setup_sheet(name):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)

# Styles
header_font = Font(bold=True)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
currency_fmt = '#,##0'

# 1. SHEET: Dau_vao
ws_dv = setup_sheet("Dau_vao")
ws_dv.column_dimensions['A'].width = 30
ws_dv.column_dimensions['B'].width = 25
ws_dv.column_dimensions['C'].width = 20
ws_dv.column_dimensions['D'].width = 20
ws_dv.column_dimensions['E'].width = 20

# Section 1: Thông số cố định
ws_dv["A1"] = "Thông số cố định"
ws_dv["A1"].font = header_font
fixed_params = [
    ("Vốn ban đầu", 1000000000),
    ("Tỷ lệ tăng trưởng tháng", 0.05),
    ("Số tháng mô phỏng", 12)
]
for i, (label, val) in enumerate(fixed_params, 2):
    ws_dv[f"A{i}"] = label
    ws_dv[f"B{i}"] = val
    ws_dv[f"B{i}"].fill = input_fill
    ws_dv[f"B{i}"].border = border
    if "Tỷ lệ" in label:
        ws_dv[f"B{i}"].number_format = '0%'
    else:
        ws_dv[f"B{i}"].number_format = currency_fmt

# Section 2, 3, 4: Cơ cấu nhân sự và Lương/Thưởng
ws_dv["A6"] = "Cơ cấu nhân sự và Định mức lương"
ws_dv["A6"].font = header_font
dv_headers = ["Ban", "Role", "Số người ban đầu", "Lương cơ bản", "Thưởng KPI"]
for c, h in enumerate(dv_headers, 1):
    cell = ws_dv.cell(row=7, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border

roles_data = [
    ("Marketing", "Manager", 1, 30000000, 5000000),
    ("Marketing", "Performance", 2, 15000000, 3000000),
    ("Marketing", "Content", 2, 12000000, 2000000),
    ("Marketing", "Creative", 1, 14000000, 2500000),
    ("Sales", "Teamlead", 1, 20000000, 10000000),
    ("Sales", "Defence", 2, 10000000, 5000000),
    ("Sales", "Member", 5, 8000000, 7000000),
    ("Accountant", "Accountant", 1, 15000000, 1000000),
    ("Tech", "Automation", 1, 25000000, 3000000),
    ("Tech", "Fullstack", 2, 35000000, 5000000),
    ("Tech", "DevOps", 1, 30000000, 4000000),
    ("OPS", "Quản lý vận hành", 1, 25000000, 4000000),
    ("OPS", "Triển khai / Delivery", 2, 15000000, 3000000),
    ("OPS", "Kiểm soát & tối ưu", 1, 18000000, 2500000),
    ("OPS", "Hỗ trợ vận hành", 2, 10000000, 1500000)
]

for r, data in enumerate(roles_data, 8):
    for c, val in enumerate(data, 1):
        cell = ws_dv.cell(row=r, column=c, value=val)
        cell.border = border
        if c >= 3:
            cell.fill = input_fill
            if c >= 4:
                cell.number_format = currency_fmt

# 2. SHEET: Luong_theo_ban
ws_l = setup_sheet("Luong_theo_ban")
l_headers = ["Tháng", "Ban", "Role", "Số người", "Lương cơ bản", "Thưởng KPI", "Tổng lương role", "Tổng lương ban"]
for c, h in enumerate(l_headers, 1):
    cell = ws_l.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

row_idx = 2
for m in range(1, 13):
    for r_idx, (ban, role, count, base, kpi) in enumerate(roles_data):
        ws_l.cell(row=row_idx, column=1, value=m).alignment = Alignment(horizontal='center')
        ws_l.cell(row=row_idx, column=2, value=ban)
        ws_l.cell(row=row_idx, column=3, value=role)
        
        # Số người (editable, reference from Dau_vao)
        c_cell = ws_l.cell(row=row_idx, column=4, value=f"=Dau_vao!$C${r_idx+8}")
        c_cell.fill = input_fill
        
        # Reference Dau_vao for salaries
        ws_l.cell(row=row_idx, column=5, value=f"=Dau_vao!$D${r_idx+8}").number_format = currency_fmt
        ws_l.cell(row=row_idx, column=6, value=f"=Dau_vao!$E${r_idx+8}").number_format = currency_fmt
        
        # Formulas
        ws_l.cell(row=row_idx, column=7, value=f"=D{row_idx}*(E{row_idx}+F{row_idx})").number_format = currency_fmt
        ws_l.cell(row=row_idx, column=8, value=f"=SUMIFS(G:G, A:A, A{row_idx}, B:B, B{row_idx})").number_format = currency_fmt
        
        for c in range(1, 9):
            ws_l.cell(row=row_idx, column=c).border = border
        row_idx += 1

ws_l.freeze_panes = "A2"
col_widths_l = {'A': 8, 'B': 15, 'C': 25, 'D': 12, 'E': 15, 'F': 15, 'G': 18, 'H': 18}
for k, v in col_widths_l.items():
    ws_l.column_dimensions[k].width = v

# 3. SHEET: KPI_theo_ban
ws_k = setup_sheet("KPI_theo_ban")
k_headers = ["Tháng", "Ban", "Tổng số người", "Ngân sách lương", "Khối lượng công việc tối thiểu", "KPI chung", "Ghi chú"]
for c, h in enumerate(k_headers, 1):
    cell = ws_k.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

depts = ["Marketing", "Sales", "Accountant", "Tech", "OPS"]
row_idx = 2
for m in range(1, 13):
    for dept in depts:
        ws_k.cell(row=row_idx, column=1, value=m).alignment = Alignment(horizontal='center')
        ws_k.cell(row=row_idx, column=2, value=dept)
        
        # Formulas from Luong_theo_ban
        ws_k.cell(row=row_idx, column=3, value=f"=SUMIFS(Luong_theo_ban!D:D, Luong_theo_ban!A:A, A{row_idx}, Luong_theo_ban!B:B, B{row_idx})")
        ws_k.cell(row=row_idx, column=4, value=f"=SUMIFS(Luong_theo_ban!G:G, Luong_theo_ban!A:A, A{row_idx}, Luong_theo_ban!B:B, B{row_idx})").number_format = currency_fmt
        
        # Editable cells
        for c in range(5, 8):
            ws_k.cell(row=row_idx, column=c).fill = input_fill
        
        for c in range(1, 8):
            ws_k.cell(row=row_idx, column=c).border = border
        row_idx += 1

ws_k.freeze_panes = "A2"
col_widths_k = {'A': 8, 'B': 15, 'C': 15, 'D': 20, 'E': 30, 'F': 15, 'G': 30}
for k, v in col_widths_k.items():
    ws_k.column_dimensions[k].width = v

wb.save(file_path)
