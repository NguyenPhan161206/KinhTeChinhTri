import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def generate_financial_model():
    wb = openpyxl.Workbook()
    
    # --- Sheet 2: Assumptions ---
    ws_assump = wb.active
    ws_assump.title = "Assumptions"
    
    assumptions = [
        ("Tham số", "Giá trị"),
        ("Vốn đầu tư ban đầu (VND)", 2000000000),
        ("Tỷ lệ tăng trưởng hàng tháng", 0.05),
        ("Lương bình quân (VND/người)", 10000000),
        ("Số nhân sự ban đầu", 30),
        ("Tuyển dụng hàng tháng", 1)
    ]
    
    for row in assumptions:
        ws_assump.append(row)
    
    # Formatting Assumptions
    for cell in ws_assump["1:1"]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    ws_assump.column_dimensions['A'].width = 30
    ws_assump.column_dimensions['B'].width = 20
    ws_assump["B3"].number_format = '0%'
    
    # Format Currency/Number cells
    ws_assump["B2"].number_format = '#,##0'
    ws_assump["B4"].number_format = '#,##0'

    # --- Sheet 1: Financial Model ---
    ws_model = wb.create_sheet("Financial Model", 0) # Insert at index 0
    
    headers = ["Tháng", "Số nhân sự", "GTDN", "Lãi tháng", "Tổng lương", "TNTT"]
    ws_model.append(headers)
    
    # Month 0 (Starting point)
    # Row index 2
    ws_model.cell(row=2, column=1, value=0) # Tháng
    ws_model.cell(row=2, column=2, value="=Assumptions!$B$5") # Số nhân sự
    ws_model.cell(row=2, column=3, value="=Assumptions!$B$2") # GTDN
    ws_model.cell(row=2, column=4, value=0) # Lãi tháng
    ws_model.cell(row=2, column=5, value=0) # Tổng lương
    ws_model.cell(row=2, column=6, value="=C2") # TNTT initial
    
    # Months 1 to 12
    for month in range(1, 13):
        row_idx = month + 2
        prev_row = row_idx - 1
        
        # A: Tháng
        ws_model.cell(row=row_idx, column=1, value=month)
        
        # B: Số nhân sự = Prev + Hiring
        ws_model.cell(row=row_idx, column=2, value=f"=B{prev_row}+Assumptions!$B$6")
        
        # C: GTDN = Prev GTDN * (1 + Growth)
        ws_model.cell(row=row_idx, column=3, value=f"=C{prev_row}*(1+Assumptions!$B$3)")
        
        # D: Lãi tháng = Prev GTDN * Growth
        ws_model.cell(row=row_idx, column=4, value=f"=C{prev_row}*Assumptions!$B$3")
        
        # E: Tổng lương = Số nhân sự * Lương bình quân
        ws_model.cell(row=row_idx, column=5, value=f"=B{row_idx}*Assumptions!$B$4")
        
        # F: TNTT_n = TNTT_(n-1) + Lai_n - Salary_n
        ws_model.cell(row=row_idx, column=6, value=f"=F{prev_row}+D{row_idx}-E{row_idx}")

    # Formatting Financial Model
    for cell in ws_model["1:1"]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    column_widths = [10, 15, 20, 20, 20, 20]
    for i, width in enumerate(column_widths, 1):
        ws_model.column_dimensions[get_column_letter(i)].width = width
        
    for row in ws_model.iter_rows(min_row=2, max_row=14, min_col=3, max_col=6):
        for cell in row:
            cell.number_format = '#,##0'

    wb.save("financial_model.xlsx")

if __name__ == "__main__":
    generate_financial_model()
