import openpyxl
import json

def extract_kpi_to_js():
    file_path = 'KTCT (1).xlsx'
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    kpi_data = {}
    # Các ban theo thứ tự xuất hiện trong Excel
    depts = ['Marketing', 'Sales', 'Tech', 'Ops']
    
    for m in range(1, 13):
        sheet_name = f'Tháng {m}'
        if sheet_name not in wb.sheetnames:
            continue
            
        ws = wb[sheet_name]
        month_data = { "Marketing": [], "Sales": [], "Tech": [], "Ops": [] }
        current_dept_idx = -1
        
        # Duyệt qua các dòng để tìm KPI
        for row in range(1, 50):
            val = ws.cell(row=row, column=1).value
            
            # Nếu gặp chữ 'Vị trí', nghĩa là bắt đầu 1 ban mới
            if val == 'Vị trí':
                current_dept_idx += 1
                continue
                
            # Nếu đang ở trong 1 ban và ô Vị trí không rỗng
            if 0 <= current_dept_idx < 4 and val is not None:
                kpi_val = ws.cell(row=row, column=4).value
                if kpi_val:
                    month_data[depts[current_dept_idx]].append({
                        "position": str(val),
                        "kpi": str(kpi_val).replace("\n", "\\n")
                    })
        
        kpi_data[m] = month_data

    # Ghi vào file JS
    with open('assets/js/data/kpiData.js', 'w', encoding='utf-8') as f:
        f.write("window.kpiData = ")
        f.write(json.dumps(kpi_data, ensure_ascii=False, indent=4))
        f.write(";")
    
    print("Successfully synchronized 12 months KPI from Excel to JS.")

if __name__ == "__main__":
    extract_kpi_to_js()
